#!/usr/bin/env python3
from __future__ import annotations

import argparse
import io
import json
import re
import urllib.request
import zipfile
from pathlib import Path

ROOT=Path(__file__).resolve().parents[1]
CFG=ROOT/'config/universal-work-ai-bot-factory.json'
OUT=ROOT/'config/generated/universal-work-ai-catalog.json'
DEFAULT_URL='https://www.onetcenter.org/dl_files/database/db_30_3_excel.zip'


def slugify(v:str)->str:
    return re.sub(r'[^a-z0-9]+','-',v.lower()).strip('-')


def classify_task(text:str)->dict:
    lower=text.lower()
    physical=any(k in lower for k in ['lift ','repair ','install ','operate machinery','drive ','construct ','clean ','cook ','assemble ','inspect equipment'])
    licensed=any(k in lower for k in ['diagnose','prescribe','represent clients in court','surgery','medical treatment','sign off engineering'])
    digital=any(k in lower for k in ['analy','research','write','prepare','calculate','record','review','schedule','communicat','design','develop','program','document','report','estimate','monitor','coordinate','plan','create'])
    if licensed: level='human_review_required'
    elif physical and not digital: level='assist_only'
    elif digital and not physical: level='sandbox_autonomous'
    elif digital: level='human_review_required'
    else: level='assist_only'
    return {'automation_level':level,'physical_world_requirement':physical,'licensed_or_high_risk_hint':licensed,'digital_work_hint':digital}


def workbook_rows(blob:bytes, name_hint:str):
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit('openpyxl required for O*NET Excel ingestion; install with: python -m pip install openpyxl') from exc
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        names=[n for n in z.namelist() if name_hint.lower() in Path(n).stem.lower() and n.lower().endswith('.xlsx')]
        if not names: raise SystemExit(f'Could not find {name_hint}.xlsx in O*NET zip')
        data=z.read(names[0])
    wb=openpyxl.load_workbook(io.BytesIO(data),read_only=True,data_only=True)
    ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); headers=[str(x).strip() if x is not None else '' for x in next(rows)]
    for values in rows:
        yield {headers[i]:values[i] for i in range(min(len(headers),len(values)))}


def main()->int:
    ap=argparse.ArgumentParser(); ap.add_argument('--onet-zip-url',default=DEFAULT_URL); ap.add_argument('--offline',action='store_true'); args=ap.parse_args()
    cfg=json.loads(CFG.read_text())
    occupations=[]; tasks=[]; source_status='not_loaded_offline'
    if not args.offline:
        with urllib.request.urlopen(args.onet_zip_url,timeout=120) as r: blob=r.read()
        occupation_rows=list(workbook_rows(blob,'Occupation Data'))
        task_rows=list(workbook_rows(blob,'Task Statements'))
        for row in occupation_rows:
            code=str(row.get('O*NET-SOC Code') or row.get('O*NET-SOC Code ') or '').strip(); title=str(row.get('Title') or '').strip(); desc=str(row.get('Description') or '').strip()
            if code and title: occupations.append({'occupation_id':code,'title':title,'description':desc,'worker_slug':f"occupation-{slugify(code)}-{slugify(title)}"})
        for row in task_rows:
            code=str(row.get('O*NET-SOC Code') or '').strip(); task_id=str(row.get('Task ID') or '').strip(); text=str(row.get('Task') or '').strip(); task_type=str(row.get('Task Type') or '').strip()
            if code and task_id and text:
                classification=classify_task(text)
                tasks.append({'task_id':f'onet-{task_id}','occupation_id':code,'task':text,'task_type':task_type,**classification,
                    'benchmark_dimensions':cfg['benchmark_dimensions'],'platform_classes':cfg['opportunity_platform_classes'],'generated_worker_status':'sandbox_only_until_graduated'})
        source_status='loaded'
    by_occ={o['occupation_id']:o for o in occupations}
    for task in tasks:
        occ=by_occ.get(task['occupation_id']); task['occupation_title']=occ['title'] if occ else None; task['worker_slug']=f"task-{slugify(task['task_id'])}-{slugify((occ or {}).get('title','work'))}"
    payload={
      'schema':'dreamco.universal_work_ai_catalog.v1','source':{'name':'O*NET','url':args.onet_zip_url,'status':source_status,'license_note':'O*NET Database available under Creative Commons terms; preserve required attribution.'},
      'occupation_count':len(occupations),'task_count':len(tasks),'occupations':occupations,'tasks':tasks,
      'automation_level_counts':{level:sum(1 for t in tasks if t['automation_level']==level) for level in cfg['automation_levels']},
      'worker_factory':cfg['worker_factory'],'money_experiment':cfg['money_experiment'],'truth_boundary':cfg['truth_rule']
    }
    OUT.parent.mkdir(parents=True,exist_ok=True); OUT.write_text(json.dumps(payload,indent=2)+'\n')
    print(json.dumps({'ok':True,'source_status':source_status,'occupations':len(occupations),'tasks':len(tasks),'output':str(OUT.relative_to(ROOT))},indent=2))
    return 0

if __name__=='__main__': raise SystemExit(main())
